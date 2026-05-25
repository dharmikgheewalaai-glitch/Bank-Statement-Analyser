import streamlit as st
import pandas as pd
import re
import io
from extractor import process_file
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, numbers

st.set_page_config(page_title="Bank Statement Analyser", layout="wide")
st.title("🏦 Bank Statement Analyser")

uploaded_file = st.file_uploader("Upload Bank Statement (PDF)", type=["pdf"])


# ------------- DATE CLEANING -------------
def clean_date(value):
    if not value:
        return None
    text = str(value).strip()
    m = re.search(r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})", text)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    return text  # return as-is rather than None so row not lost


# ------------- AMOUNT CLEANING -------------
def clean_amount(value):
    if value is None or str(value).strip() in ("", "None"):
        return 0.00
    try:
        return round(float(str(value).replace(",", "").strip()), 2)
    except (ValueError, TypeError):
        return 0.00


# ------------- SUMMARY STATS -------------
def show_summary(df):
    total_debit  = df["Debit"].sum()
    total_credit = df["Credit"].sum()
    net          = total_credit - total_debit

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Debits",  f"₹ {total_debit:,.2f}")
    c2.metric("Total Credits", f"₹ {total_credit:,.2f}")
    c3.metric("Net Flow",      f"₹ {net:,.2f}", delta=f"{net:,.2f}")

    st.markdown("---")

    # Monthly Head-wise breakdown
    with st.expander("📊 Monthly Head-wise Breakdown"):
        df2 = df.copy()

        # Parse Month from Date (DD/MM/YYYY)
        def extract_month(d):
            m = re.match(r"\d{1,2}/(\d{1,2})/(\d{4})", str(d or ""))
            if m:
                mo, yr = int(m.group(1)), m.group(2)
                return f"{yr}-{mo:02d}"
            return "Unknown"

        df2["Month"] = df2["Date"].apply(extract_month)

        months = sorted(df2["Month"].unique())
        sel_month = st.selectbox("Select Month", ["All"] + months, key="month_sel")

        if sel_month != "All":
            df2 = df2[df2["Month"] == sel_month]

        hw_raw = df2.groupby("Head").agg(
            Debit=("Debit", "sum"),
            Credit=("Credit", "sum"),
            Count=("Date", "count"),
        ).reset_index().sort_values("Debit", ascending=False)

        hw = hw_raw.copy()
        hw["Debit"]  = hw["Debit"].map(lambda x: f"{x:,.2f}")
        hw["Credit"] = hw["Credit"].map(lambda x: f"{x:,.2f}")
        st.dataframe(hw, use_container_width=True)

        label = sel_month if sel_month != "All" else "all"
        dl_csv = hw_raw.to_csv(index=False, float_format="%.2f").encode("utf-8")

        hw_xl = io.BytesIO()
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Monthly Head-wise"
        for row in dataframe_to_rows(hw_raw, index=False, header=True):
            ws2.append(row)
        for col_cells in ws2.columns:
            ws2.column_dimensions[col_cells[0].column_letter].width = max(
                len(str(c.value or "")) for c in col_cells) + 2
        wb2.save(hw_xl)
        hw_xl.seek(0)

        dc1, dc2 = st.columns(2)
        with dc1:
            st.download_button("⬇️ Download CSV",
                               data=dl_csv,
                               file_name=f"headwise_{label}.csv",
                               key="hw_csv")
        with dc2:
            st.download_button("⬇️ Download Excel",
                               data=hw_xl,
                               file_name=f"headwise_{label}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="hw_xl")


# ===================================================================
#                           MAIN
# ===================================================================
if uploaded_file is not None:
    st.info(f"Processing: **{uploaded_file.name}** …")
    file_bytes = uploaded_file.read()

    meta, transactions = process_file(file_bytes, uploaded_file.name)

    # Show any parse warnings
    if meta.get("_logs"):
        with st.expander("⚠️ Parser Warnings"):
            for log in meta["_logs"]:
                st.text(log)

    if not transactions:
        st.error("No transactions found. Check that PDF has selectable text.")
        st.stop()

    df = pd.DataFrame(transactions)

    # ---------- CLEAN ----------
    df_final = df.copy()

    if "Date" in df_final.columns:
        df_final["Date"] = df_final["Date"].apply(clean_date)

    for col in ["Debit", "Credit", "Balance"]:
        if col in df_final.columns:
            df_final[col] = df_final[col].apply(clean_amount)

    # Reorder: put Head after Particulars, Balance last
    desired_order = ["Page", "Date", "Particulars", "Debit", "Credit", "Head", "Balance"]
    cols = [c for c in desired_order if c in df_final.columns]
    extra = [c for c in df_final.columns if c not in cols]
    df_final = df_final[cols + extra]

    # Drop Page col from display/export (keep for debug if needed)
    df_export = df_final.drop(columns=["Page"], errors="ignore")

    # ---------- SUMMARY ----------
    show_summary(df_export)

    # ---------- DISPLAY ----------
    df_display = df_export.copy()
    for col in ["Debit", "Credit", "Balance"]:
        if col in df_display.columns:
            df_display[col] = df_display[col].map(lambda x: f"{x:,.2f}")

    st.success(f"✅ {len(df_export)} transactions extracted.")
    st.dataframe(df_display, use_container_width=True)

    # ---------- BASE NAMES ----------
    base = uploaded_file.name.rsplit(".", 1)[0]

    # ---------- CSV ----------
    csv_bytes = df_export.to_csv(index=False, float_format="%.2f").encode("utf-8")

    # ---------- EXCEL ----------
    excel_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), start=1):
        ws.append(row)
        for c_idx, cell in enumerate(ws[r_idx], start=1):
            hdr = ws.cell(row=1, column=c_idx).value
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            elif hdr in ("Debit", "Credit", "Balance"):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Auto-width columns
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(length, 50)

    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # ---------- PDF ----------
    df_pdf = df_export.copy()
    for col in ["Debit", "Credit", "Balance"]:
        if col in df_pdf.columns:
            df_pdf[col] = df_pdf[col].map(lambda x: f"{x:,.2f}")

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Bank Statement — {base}", styles["Title"]),
        Spacer(1, 12),
    ]

    # Wrap long particulars text
    def wrap(text, max_chars=40):
        return str(text) if len(str(text)) <= max_chars else str(text)[:max_chars] + "…"

    data_rows = [list(df_pdf.columns)]
    for _, row in df_pdf.iterrows():
        data_rows.append([
            wrap(v, 40) if col == "Particulars" else str(v)
            for col, v in zip(df_pdf.columns, row)
        ])

    col_widths = []
    for col in df_pdf.columns:
        if col == "Particulars":
            col_widths.append(200)
        elif col in ("Debit", "Credit", "Balance"):
            col_widths.append(70)
        elif col == "Date":
            col_widths.append(65)
        elif col == "Head":
            col_widths.append(90)
        else:
            col_widths.append(50)

    table = Table(data_rows, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("ALIGN",       (0, 1), (-1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.grey),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)
    pdf_buffer.seek(0)

    # ---------- DOWNLOADS ----------
    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        st.download_button("⬇️ Download CSV",   data=csv_bytes,      file_name=f"{base}.csv")
    with dl2:
        st.download_button("⬇️ Download Excel", data=excel_buffer,   file_name=f"{base}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with dl3:
        st.download_button("⬇️ Download PDF",   data=pdf_buffer,     file_name=f"{base}.pdf",
                           mime="application/pdf")
